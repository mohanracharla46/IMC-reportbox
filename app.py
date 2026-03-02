"""
Work Report Management System
A Flask-based application for managing daily employee work reports
"""

from flask import Flask, render_template, request, redirect, url_for, flash, session, send_from_directory
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from functools import wraps
from datetime import datetime, date
import sqlite3
import psycopg2
from psycopg2.extras import RealDictCursor
import os
import pandas as pd
from io import BytesIO
import urllib.parse as urlparse
import re

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-in-production'  # Change this in production!
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
ALLOWED_EXTENSIONS = {'txt', 'pdf', 'doc', 'docx', 'xls', 'xlsx', 'png', 'jpg', 'jpeg', 'gif'}

# Ensure upload directory exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

@app.errorhandler(413)
def request_entity_too_large(error):
    flash('File is too large! Maximum size allowed is 16MB.', 'error')
    return redirect(request.referrer or url_for('index'))

@app.errorhandler(404)
def page_not_found(error):
    return render_template('login.html'), 404

@app.errorhandler(500)
def internal_server_error(error):
    return "Internal Server Error", 500

def calculate_submission_amount(work_type, quantity, employment_type):
    """Calculate amount based on work type for both inhouse and freelancers"""
    
    # Pricing structure (applies to both inhouse and freelancers)
    rates = {
        'Shoot with Camera': 2000,
        'Calendar Poster': 300,
        'Informative Poster': 300,
        'Elevation Poster': 300,
        'Reel': 600,
        'Press Conference Shoot': 400,
        'Regular Video': 1500,
        'Cinematic Videos': 2500,
        'Brochures': 1000,
        'Printing Material': 500,
        'Shoot on Mobile': 2500,
        'Poster': 300,
        'Video': 1500,
        'Brochure': 1000,
        'Print Material': 500,
        'Other': 0
    }
    
    rate = rates.get(work_type, 0)
    
    try:
        qty = int(quantity or 0)
    except (ValueError, TypeError):
        qty = 0
        
    return rate * qty

def format_datetime_filter(value, only_time=False):
    """Safe formatting for created_at (works for string and datetime)"""
    if not value:
        return ""
    
    # If it's already a string (SQLite)
    if isinstance(value, str):
        # Handle formats like 2024-02-15 13:01:17.123456 or 2024-02-15 13:01:17
        try:
            # Remove milliseconds if present
            val = value.split('.')[0]
            if only_time and ' ' in val:
                return val.split(' ')[1]
            return val
        except:
            return value
        
    # If it's a datetime/date object (Postgres)
    if hasattr(value, 'strftime'):
        try:
            if only_time:
                return value.strftime("%H:%M:%S")
            return value.strftime("%Y-%m-%d %H:%M:%S")
        except:
            return str(value)
        
    return str(value)

def get_filename_filter(file_path):
    """Extract filename from path safely"""
    if not file_path:
        return ""
    return os.path.basename(file_path)

app.jinja_env.globals.update(calculate_submission_amount=calculate_submission_amount)
app.jinja_env.filters['format_dt'] = format_datetime_filter
app.jinja_env.filters['filename'] = get_filename_filter

# Database helper
def format_streak_dates(raw_data):
    """Ensure all dates returned for streaks are safely formatted as YYYY-MM-DD strings to prevent JSON serialization issues."""
    formatted = []
    for r in raw_data:
        d = dict(r)
        date_val = d.get('work_date')
        if hasattr(date_val, 'strftime'):
            d['work_date'] = date_val.strftime('%Y-%m-%d')
        elif isinstance(date_val, str):
            # Fallback string parsing just in case DB returned something unusual natively
            d['work_date'] = str(date_val).split(' ')[0].replace('/', '-')
        formatted.append(d)
    return formatted

def get_db_info():
    """Get database connection info and placeholder type"""
    db_url = os.environ.get('DATABASE_URL')
    if db_url:
        if db_url.startswith("postgres://"):
            db_url = db_url.replace("postgres://", "postgresql://", 1)
        return db_url, '%s'
    return 'work_reports.db', '?'

# Database initialization
def init_db():
    """Initialize the database with required tables"""
    db_url, q = get_db_info()
    
    if q == '%s':
        conn = psycopg2.connect(db_url)
        conn.autocommit = True
        cursor = conn.cursor()
        id_type = "SERIAL PRIMARY KEY"
        check_constraint_role = "CHECK(role IN ('employee', 'admin'))"
        check_constraint_type = "CHECK(employment_type IN ('inhouse', 'freelancer'))"
    else:
        conn = sqlite3.connect(db_url)
        cursor = conn.cursor()
        id_type = "INTEGER PRIMARY KEY AUTOINCREMENT"
        check_constraint_role = "CHECK(role IN ('employee', 'admin'))"
        check_constraint_type = "CHECK(employment_type IN ('inhouse', 'freelancer'))"

    # Create users table
    cursor.execute(f'''
        CREATE TABLE IF NOT EXISTS users (
            id {id_type},
            name TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT NOT NULL {check_constraint_role},
            employment_type TEXT DEFAULT 'inhouse' {check_constraint_type}
        )
    ''')
    
    # Create submissions table
    cursor.execute(f'''
        CREATE TABLE IF NOT EXISTS submissions (
            id {id_type},
            user_id INTEGER NOT NULL,
            work_text TEXT NOT NULL,
            client_category TEXT,
            client_name TEXT,
            work_type TEXT,
            quantity INTEGER DEFAULT 1,
            file_path TEXT,
            date DATE NOT NULL,
            submission_number INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
        )
    ''')
    
    # Add columns to existing tables if they don't exist
    for column, table, definition in [
        ('employment_type', 'users', 'TEXT DEFAULT "inhouse"'),
        ('client_category', 'submissions', 'TEXT'),
        ('client_name', 'submissions', 'TEXT'),
        ('work_type', 'submissions', 'TEXT'),
        ('quantity', 'submissions', 'INTEGER DEFAULT 1'),
        ('employee_name', 'submissions', 'TEXT')
    ]:
        try:
            cursor.execute(f'ALTER TABLE {table} ADD COLUMN {column} {definition}')
        except (sqlite3.OperationalError, psycopg2.Error):
            pass # Usually means column already exists
    
    # Create default admin if not exists
    cursor.execute(f'SELECT * FROM users WHERE role = {q}', ('admin',))
    if not cursor.fetchone():
        admin_password = generate_password_hash('admin123')
        cursor.execute(
            f'INSERT INTO users (name, email, password, role, employment_type) VALUES ({q}, {q}, {q}, {q}, {q})',
            ('Prashanth', 'prashanth@iramediaconcepts.com', admin_password, 'admin', 'inhouse')
        )
    else:
        cursor.execute(
            f'UPDATE users SET email = {q}, name = {q} WHERE role = {q} AND email = {q}',
            ('prashanth@iramediaconcepts.com', 'Prashanth', 'admin', 'admin@company.com')
        )
    
    conn.commit()
    conn.close()

# Initialize database on startup
init_db()

def get_db_connection():
    """Create and return a database connection"""
    db_url, q = get_db_info()
    if q == '%s':
        conn = psycopg2.connect(db_url)
        conn.autocommit = True
        return conn
    else:
        conn = sqlite3.connect(db_url)
        conn.row_factory = sqlite3.Row
        return conn

def execute_query(conn, query, params=None):
    """Helper to execute query correctly regardless of DB type"""
    _, q = get_db_info()
    if q == '%s':
        query = query.replace('?', '%s')
        # Robust strftime translation
        if 'strftime' in query:
            # Replaces strftime('%Y-%m', col) or strftime("%Y-%m", col) with TO_CHAR(col, 'YYYY-MM')
            # Added \s* around params to handle formatting differences
            query = re.sub(r"strftime\(\s*['\"]%Y-%m['\"]\s*,\s*(\w+\.?\w+)\s*\)", r"TO_CHAR(\1, 'YYYY-MM')", query)
        
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute(query, params or ())
        return cursor
    else:
        return conn.execute(query, params or ())

def allowed_file(filename):
    """Check if uploaded file has an allowed extension"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Decorator for login required
def login_required(f):
    """Decorator to ensure user is logged in"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Decorator for admin only
def admin_required(f):
    """Decorator to ensure user is an admin"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'error')
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            flash('Access denied. Admin privileges required.', 'error')
            return redirect(url_for('employee_dashboard'))
        return f(*args, **kwargs)
    return decorated_function

# Routes
@app.route('/')
def index():
    """Redirect to appropriate dashboard based on login status"""
    if 'user_id' in session:
        if session.get('role') == 'admin':
            return redirect(url_for('admin_dashboard'))
        return redirect(url_for('employee_dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Handle user login"""
    if request.method == 'POST':
        name = request.form.get('name')
        password = request.form.get('password')
        
        if not name or not password:
            flash('Please provide both name and password.', 'error')
            return render_template('login.html')
        
        conn = get_db_connection()
        # Case-insensitive login
        user = execute_query(conn, 'SELECT * FROM users WHERE LOWER(name) = LOWER(?)', (name.strip(),)).fetchone()
        conn.close()
        
        if not user:
            flash('No user found with this name.', 'error')
        elif not check_password_hash(user['password'], password):
            flash('Incorrect password. Please try again.', 'error')
        else:
            session['user_id'] = user['id']
            session['user_name'] = user['name']
            session['role'] = user['role']
            session['employment_type'] = dict(user).get('employment_type', 'inhouse')
            flash(f'Welcome back, {user["name"]}!', 'success')
            
            if user['role'] == 'admin':
                return redirect(url_for('admin_dashboard'))
            return redirect(url_for('employee_dashboard'))
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    """Handle user logout"""
    name = session.get('user_name', 'User')
    session.clear()
    flash(f'Goodbye, {name}! You have been logged out.', 'success')
    return redirect(url_for('login'))

@app.route('/employee/dashboard')
@login_required
def employee_dashboard():
    """Employee dashboard - view submission status and submit reports"""
    if session.get('role') == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    conn = get_db_connection()
    
    # Check if submitted today and count submissions
    today = date.today().isoformat()
    submissions_today = execute_query(conn, 
        'SELECT * FROM submissions WHERE user_id = ? AND date = ? ORDER BY created_at DESC',
        (session['user_id'], today)
    ).fetchall()
    
    submission_count_today = len(submissions_today)
    
    # Get recent submissions (filtered or last 20)
    date_filter = request.args.get('date')
    if date_filter:
        recent_submissions = [dict(row) for row in execute_query(conn, 
            'SELECT * FROM submissions WHERE user_id = ? AND date = ? ORDER BY created_at DESC',
            (session['user_id'], date_filter)
        ).fetchall()]
    else:
        recent_submissions = [dict(row) for row in execute_query(conn, 
            'SELECT * FROM submissions WHERE user_id = ? ORDER BY date DESC, created_at DESC LIMIT 20',
            (session['user_id'],)
        ).fetchall()]
    
    # Get streak data for the user
    raw_streak_results = execute_query(conn,
        '''SELECT date as work_date, SUM(CAST(quantity AS INTEGER)) as daily_qty
           FROM submissions
           WHERE user_id = ?
           GROUP BY date ORDER BY date ASC''',
        (session['user_id'],)
    ).fetchall()
    
    raw_streak = format_streak_dates(raw_streak_results)
    
    # Get client distribution for the user
    client_distribution_results = execute_query(conn,
        '''SELECT client_name, SUM(CAST(quantity AS INTEGER)) as total_qty
           FROM submissions
           WHERE user_id = ?
           GROUP BY client_name''',
        (session['user_id'],)
    ).fetchall()
    
    client_distribution = [dict(row) for row in client_distribution_results]
    
    conn.close()
    
    return render_template(
        'employee_dashboard.html',
        submitted_today=submission_count_today > 0,
        submission_count_today=submission_count_today,
        submissions_today=submissions_today,
        recent_submissions=recent_submissions,
        raw_streak=raw_streak,
        client_distribution=client_distribution,
        now=datetime.now()
    )

@app.route('/employee/submit', methods=['POST'])
@login_required
def submit_report():
    """Handle work report submission - multiple submissions allowed per day"""
    if session.get('role') == 'admin':
        flash('Admins cannot submit work reports.', 'error')
        return redirect(url_for('admin_dashboard'))
    
    work_text = request.form.get('work_text', '').strip()
    client_category = request.form.get('client_category', '').strip()
    client_name = request.form.get('client_name', '').strip()
    other_client_name = request.form.get('other_client_name', '').strip()
    work_type = request.form.get('work_type', '').strip()
    other_work_type_name = request.form.get('other_work_type_name', '').strip()
    quantity = request.form.get('quantity', 1)
    
    today = date.today().isoformat()
    
    # Allow backdating if provided
    submission_date = request.form.get('submission_date')
    if submission_date:
        try:
            submission_dt = datetime.strptime(submission_date, '%Y-%m-%d').date()
            if submission_dt > date.today():
                flash('Cannot submit reports for future dates.', 'error')
                return redirect(url_for('employee_dashboard'))
            today = submission_date
        except ValueError:
            pass # Keep today as default if invalid format
    
    # Handle 'Other' client name
    if client_name in ['Others', 'Other'] and other_client_name:
        client_name = other_client_name

    # Handle 'Other' work type
    if work_type == 'Other' and other_work_type_name:
        work_type = other_work_type_name

    if not work_text or not client_name:
        flash('Please provide work description and client name.', 'error')
        return redirect(url_for('employee_dashboard'))
    
    # Handle file upload
    file_path = None
    if 'file' in request.files:
        file = request.files['file']
        if file and file.filename and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            # Add timestamp to filename to avoid conflicts
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{session['user_id']}_{timestamp}_{filename}"
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
    
    conn = get_db_connection()
    
    # Check for duplicate submission (same text, client, and work type on same day)
    existing_check = execute_query(conn, 
        '''SELECT id FROM submissions 
           WHERE user_id = ? AND date = ? AND work_text = ? AND client_name = ? AND work_type = ?''',
        (session['user_id'], today, work_text, client_name, work_type)
    ).fetchone()
    
    if existing_check:
        conn.close()
        flash('You have already submitted an identical report today.', 'info')
        return redirect(url_for('employee_dashboard'))
    
    # Get submission count for today to set submission_number
    res = execute_query(conn, 
        'SELECT COUNT(*) as count FROM submissions WHERE user_id = ? AND date = ?',
        (session['user_id'], today)
    ).fetchone()
    submission_count = res['count']
    
    submission_number = submission_count + 1
    
    # Insert new submission
    try:
        execute_query(conn, 
            'INSERT INTO submissions (user_id, work_text, client_category, client_name, work_type, quantity, file_path, date, submission_number, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
            (session['user_id'], work_text, client_category, client_name, work_type, quantity, file_path, today, submission_number, datetime.now())
        )
        conn.commit()
        flash(f'Work report #{submission_number} submitted successfully!', 'success')
    except Exception as e:
        flash(f'Error submitting report: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('employee_dashboard'))

@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    """Admin dashboard - view and manage all submissions"""
    conn = get_db_connection()
    
    # Get filter parameters
    employee_filter = request.args.get('employee', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    employment_type_filter = request.args.get('employment_type', 'both')
    
    # Build query
    query = '''
        SELECT s.*, 
               COALESCE(s.employee_name, u.name) as display_name,
               u.email as employee_email, 
               u.employment_type
        FROM submissions s
        JOIN users u ON s.user_id = u.id
        WHERE 1=1
    '''
    params = []
    
    if employee_filter:
        query += ' AND (LOWER(u.name) LIKE LOWER(?) OR LOWER(s.employee_name) LIKE LOWER(?))'
        params.extend([f'%{employee_filter}%', f'%{employee_filter}%'])
    
    if start_date:
        query += ' AND s.date >= ?'
        params.append(start_date)
        
    if end_date:
        query += ' AND s.date <= ?'
        params.append(end_date)
        
    if employment_type_filter and employment_type_filter != 'both':
        query += ' AND u.employment_type = ?'
        params.append(employment_type_filter)
    
    query += ' ORDER BY s.date DESC, s.created_at DESC'
    
    submissions = [dict(row) for row in execute_query(conn, query, params).fetchall()]
    
    # Get all employees and freelancers separately
    employees_raw = [dict(row) for row in execute_query(conn, 
        'SELECT * FROM users WHERE role = ? AND employment_type = ? ORDER BY name',
        ('employee', 'inhouse')
    ).fetchall()]
    
    freelancers = [dict(row) for row in execute_query(conn, 
        'SELECT * FROM users WHERE role = ? AND employment_type = ? ORDER BY name',
        ('employee', 'freelancer')
    ).fetchall()]

    # Calculate monthly amounts for both employees and freelancers
    current_month = date.today().strftime('%Y-%m')
    
    # Process employees (inhouse)
    employees_with_amounts = []
    for e in employees_raw:
        e_dict = dict(e)
        # Get all submissions for this employee for the current month
        e_submissions = execute_query(conn, 
            'SELECT work_type, quantity FROM submissions WHERE user_id = ? AND strftime("%Y-%m", date) = ?',
            (e['id'], current_month)
        ).fetchall()
        
        total_amount = sum(calculate_submission_amount(s['work_type'], s['quantity'], 'inhouse') for s in e_submissions)
        e_dict['monthly_amount'] = total_amount
        employees_with_amounts.append(e_dict)
    
    # Process freelancers
    freelancers_with_amounts = []
    for f in freelancers:
        f_dict = dict(f)
        # Get all submissions for this freelancer for the current month
        f_submissions = execute_query(conn, 
            'SELECT work_type, quantity FROM submissions WHERE user_id = ? AND strftime("%Y-%m", date) = ?',
            (f['id'], current_month)
        ).fetchall()
        
        total_amount = sum(calculate_submission_amount(s['work_type'], s['quantity'], 'freelancer') for s in f_submissions)
        f_dict['monthly_amount'] = total_amount
        freelancers_with_amounts.append(f_dict)
    
    # Get statistics
    total_submissions = execute_query(conn, 'SELECT COUNT(*) as count FROM submissions').fetchone()['count']
    total_employees = len(employees_with_amounts)
    today_submissions = execute_query(conn, 
        'SELECT COUNT(*) as count FROM submissions WHERE date = ?',
        (date.today().isoformat(),)
    ).fetchone()['count']
    
    conn.close()
    
    return render_template(
        'admin_dashboard.html',
        submissions=submissions,
        employees=employees_with_amounts,
        freelancers=freelancers_with_amounts,
        total_submissions=total_submissions,
        total_employees=len(employees_with_amounts),
        total_freelancers=len(freelancers),
        today_submissions=today_submissions,
        employee_filter=employee_filter,
        start_date=start_date,
        end_date=end_date,
        employment_type_filter=employment_type_filter,
        now=datetime.now()
    )

@app.route('/admin/submit', methods=['POST'])
@admin_required
def admin_submit_report():
    """Handle work report submission by admin for their own work (Self Work)"""
    employee_name = request.form.get('employee_name', '').strip()
    other_employee_name = request.form.get('other_employee_name', '').strip()
    # If "Other" was selected, use the manually typed name
    if employee_name == 'Other' and other_employee_name:
        employee_name = other_employee_name
    work_text = request.form.get('work_text', '').strip()
    client_category = request.form.get('client_category', '').strip()
    client_name = request.form.get('client_name', '').strip()
    other_client_name = request.form.get('other_client_name', '').strip()
    work_type = request.form.get('work_type', '').strip()
    other_work_type_name = request.form.get('other_work_type_name', '').strip()
    quantity = request.form.get('quantity', 1)
    submission_date = request.form.get('submission_date')
    
    user_id = session['user_id'] # Use the admin's own ID
    today = date.today().isoformat()
    
    if submission_date:
        try:
            submission_dt = datetime.strptime(submission_date, '%Y-%m-%d').date()
            if submission_dt > date.today():
                flash('Cannot submit reports for future dates.', 'error')
                return redirect(url_for('admin_dashboard'))
            today = submission_date
        except ValueError:
            pass
            
    # Handle 'Other' client name
    if client_name in ['Others', 'Other'] and other_client_name:
        client_name = other_client_name

    # Handle 'Other' work type
    if work_type == 'Other' and other_work_type_name:
        work_type = other_work_type_name

    if not work_text or not client_name:
        flash('Please provide work description and client name.', 'error')
        return redirect(url_for('admin_dashboard'))
    
    # Handle file upload
    file_path = None
    if 'file' in request.files:
        file = request.files['file']
        if file and file.filename and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"admin_{user_id}_{timestamp}_{filename}"
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
    
    conn = get_db_connection()
    # Check for duplicate submission
    existing_check = execute_query(conn, 
        '''SELECT id FROM submissions 
           WHERE user_id = ? AND date = ? AND work_text = ? AND client_name = ? AND work_type = ?''',
        (user_id, today, work_text, client_name, work_type)
    ).fetchone()
    
    if existing_check:
        conn.close()
        flash('An identical report already exists for today.', 'info')
        return redirect(url_for('admin_dashboard'))
    
    # Get submission count
    res = execute_query(conn, 
        'SELECT COUNT(*) as count FROM submissions WHERE user_id = ? AND date = ?',
        (user_id, today)
    ).fetchone()
    submission_number = res['count'] + 1
    
    try:
        execute_query(conn, 
            'INSERT INTO submissions (user_id, work_text, client_category, client_name, work_type, quantity, file_path, date, submission_number, created_at, employee_name) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
            (user_id, work_text, client_category, client_name, work_type, quantity, file_path, today, submission_number, datetime.now(), employee_name)
        )
        conn.commit()
        flash(f'Self Work report #{submission_number} submitted successfully!', 'success')
    except Exception as e:
        flash(f'Error submitting report: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/export-self-work')
@admin_required
def export_self_work():
    """Download reports submitted by admins (Self Work) as Excel"""
    from flask import make_response
    conn = get_db_connection()
    
    query = '''
        SELECT s.date as "Date", 
               s.employee_name as "Submitted Name",
               u.name as "Admin User",
               s.client_category as "Category",
               s.client_name as "Client",
               s.work_type as "Work Type",
               s.quantity as "Qty",
               s.work_text as "Description",
               s.created_at as "Submitted At"
        FROM submissions s
        JOIN users u ON s.user_id = u.id
        WHERE u.role = 'admin'
        ORDER BY s.date DESC, s.created_at DESC
    '''
    
    df = pd.read_sql_query(query, conn)
    conn.close()
    
    if df.empty:
        flash('No self work reports found to export.', 'info')
        return redirect(url_for('admin_dashboard'))

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Self Work Reports')
        worksheet = writer.sheets['Self Work Reports']
        for i, col in enumerate(df.columns):
            max_len = df[col].astype(str).map(len).max()
            column_len = max(max_len, len(str(col))) + 2
            col_letter = chr(65 + i)
            worksheet.column_dimensions[col_letter].width = min(column_len, 50)

    output.seek(0)
    response = make_response(output.getvalue())
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response.headers["Content-Disposition"] = f"attachment; filename=Self_Work_Export_{timestamp}.xlsx"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response

@app.route('/admin/employee/add', methods=['POST'])
@admin_required
def add_employee():
    """Add a new employee"""
    name = request.form.get('name', '').strip()
    email = request.form.get('email', '').strip()
    password = request.form.get('password', '').strip()
    employment_type = request.form.get('employment_type', 'inhouse')
    
    if not all([name, email, password]):
        flash('All fields are required.', 'error')
        return redirect(url_for('admin_dashboard'))
    
    conn = get_db_connection()
    
    
    # Check if name or email already exists
    existing_name = execute_query(conn, 'SELECT * FROM users WHERE LOWER(name) = LOWER(?)', (name,)).fetchone()
    existing_email = execute_query(conn, 'SELECT * FROM users WHERE email = ?', (email,)).fetchone()
    
    if existing_name:
        flash('Employee name already exists. Please choose a different name.', 'error')
        conn.close()
        return redirect(url_for('admin_dashboard'))

    if existing_email:
        flash('Email already exists.', 'error')
        conn.close()
        return redirect(url_for('admin_dashboard'))
    
    # Insert new employee
    try:
        hashed_password = generate_password_hash(password)
        execute_query(conn, 
            'INSERT INTO users (name, email, password, role, employment_type) VALUES (?, ?, ?, ?, ?)',
            (name, email, hashed_password, 'employee', employment_type)
        )
        conn.commit()
        flash(f'Employee {name} added successfully!', 'success')
    except Exception as e:
        flash(f'Error adding employee: {str(e)}', 'error')
    finally:
        conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/employee/edit/<int:employee_id>', methods=['POST'])
@admin_required
def edit_employee(employee_id):
    """Edit an existing employee"""
    name = request.form.get('name', '').strip()
    email = request.form.get('email', '').strip()
    password = request.form.get('password', '').strip()
    employment_type = request.form.get('employment_type', 'inhouse')
    
    if not all([name, email]):
        flash('Name and email are required.', 'error')
        return redirect(url_for('admin_dashboard'))
    
    conn = get_db_connection()
    
    # Check if employee exists
    employee = execute_query(conn, 'SELECT * FROM users WHERE id = ? AND role = ?', 
                           (employee_id, 'employee')).fetchone()
    if not employee:
        flash('Employee not found.', 'error')
        conn.close()
        return redirect(url_for('admin_dashboard'))
    
    # Check if new email is already taken by another user
    if email != employee['email']:
        existing_email = execute_query(conn, 'SELECT * FROM users WHERE email = ? AND id != ?', (email, employee_id)).fetchone()
        if existing_email:
            flash('Email already in use by another employee.', 'error')
            conn.close()
            return redirect(url_for('admin_dashboard'))

    # Check if new name is already taken by another user
    if name != employee['name']:
        existing_name = execute_query(conn, 'SELECT * FROM users WHERE LOWER(name) = LOWER(?) AND id != ?', (name, employee_id)).fetchone()
        if existing_name:
            flash('Name already in use by another employee.', 'error')
            conn.close()
            return redirect(url_for('admin_dashboard'))

    # Update employee
    try:
        if password:
            hashed_password = generate_password_hash(password)
            execute_query(conn, 
                'UPDATE users SET name = ?, email = ?, password = ?, employment_type = ? WHERE id = ?',
                (name, email, hashed_password, employment_type, employee_id)
            )
        else:
            execute_query(conn, 
                'UPDATE users SET name = ?, email = ?, employment_type = ? WHERE id = ?',
                (name, email, employment_type, employee_id)
            )
        conn.commit()
        flash(f'Employee {name} updated successfully!', 'success')
    except Exception as e:
        flash(f'Error updating employee: {str(e)}', 'error')
    finally:
        conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/employee/delete/<int:employee_id>', methods=['POST'])
@admin_required
def delete_employee(employee_id):
    """Delete an employee"""
    conn = get_db_connection()
    
    # Check if employee exists
    employee = execute_query(conn, 'SELECT * FROM users WHERE id = ? AND role = ?',
                           (employee_id, 'employee')).fetchone()
    if not employee:
        flash('Employee not found.', 'error')
        conn.close()
        return redirect(url_for('admin_dashboard'))
    
    # Delete employee (submissions will be deleted due to CASCADE)
    execute_query(conn, 'DELETE FROM users WHERE id = ?', (employee_id,))
    conn.commit()
    conn.close()
    
    flash(f'Employee {employee["name"]} deleted successfully!', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/download/<path:filename>')
@login_required
def download_file(filename):
    """Download uploaded files"""
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename, as_attachment=True)

@app.route('/admin/employee/<int:employee_id>/reports')
@admin_required
def view_employee_reports(employee_id):
    """View all reports for a specific employee"""
    conn = get_db_connection()
    
    # Get employee details
    employee = execute_query(conn, 'SELECT * FROM users WHERE id = ?', (employee_id,)).fetchone()
    if not employee:
        flash('Employee not found.', 'error')
        conn.close()
        return redirect(url_for('admin_dashboard'))
    
    # Get all submissions for this employee
    submissions = [dict(row) for row in execute_query(conn, 
        '''SELECT * FROM submissions 
           WHERE user_id = ? 
           ORDER BY date DESC, created_at DESC''',
        (employee_id,)
    ).fetchall()]
    
    # Group submissions by month for statistics
    monthly_stats_raw = execute_query(conn, 
        '''SELECT strftime('%Y-%m', date) as month, 
           COUNT(*) as count,
           COUNT(DISTINCT date) as days_submitted
           FROM submissions 
           WHERE user_id = ?
           GROUP BY month
           ORDER BY month DESC''',
        (employee_id,)
    ).fetchall()
    
    monthly_stats = []
    for stat in monthly_stats_raw:
        # Filter submissions for this month - handle both string and date object
        month_submissions = []
        for s in submissions:
            s_date = s['date'].isoformat() if hasattr(s['date'], 'isoformat') else str(s['date'])
            if s_date.startswith(stat['month']):
                month_submissions.append(s)
                
        total_amount = sum(calculate_submission_amount(s['work_type'], s['quantity'], employee['employment_type']) for s in month_submissions)
        
        monthly_stats.append({
            'month': stat['month'],
            'count': stat['count'],
            'days_submitted': stat['days_submitted'],
            'total_amount': total_amount
        })
    
    conn.close()
    
    return render_template(
        'employee_reports.html',
        employee=employee,
        submissions=submissions,
        monthly_stats=monthly_stats
    )

@app.route('/admin/employee/<int:employee_id>/monthly-report')
@admin_required
def view_monthly_report(employee_id):
    """View monthly report for an employee"""
    month = request.args.get('month')  # Format: YYYY-MM
    
    if not month:
        flash('Please select a month.', 'error')
        return redirect(url_for('view_employee_reports', employee_id=employee_id))
    
    conn = get_db_connection()
    
    # Get employee details
    employee = execute_query(conn, 'SELECT * FROM users WHERE id = ?', (employee_id,)).fetchone()
    if not employee:
        flash('Employee not found.', 'error')
        conn.close()
        return redirect(url_for('admin_dashboard'))
    
    # Get submissions for the specified month
    submissions = [dict(row) for row in execute_query(conn, 
        '''SELECT * FROM submissions 
           WHERE user_id = ? AND strftime('%Y-%m', date) = ?
           ORDER BY date ASC, created_at ASC''',
        (employee_id, month)
    ).fetchall()]
    
    # Calculate daily totals for freelancers
    daily_totals = {}
    for s in submissions:
        # Use string as key for consistency across DB types
        d = s['date'].isoformat() if hasattr(s['date'], 'isoformat') else str(s['date'])
        if d not in daily_totals:
            daily_totals[d] = 0
        daily_totals[d] += calculate_submission_amount(s['work_type'], s['quantity'], employee['employment_type'])
    # Calculate total monthly amount
    total_amount = sum(calculate_submission_amount(s['work_type'], s['quantity'], employee['employment_type']) for s in submissions)
    
    conn.close()
    
    return render_template(
        'monthly_report.html',
        employee=employee,
        submissions=submissions,
        month=month,
        total_amount=total_amount,
        daily_totals=daily_totals
    )

@app.route('/admin/employee/<int:employee_id>/monthly-report/download')
@login_required
def download_monthly_report(employee_id):
    """Download monthly report as Excel file"""
    from flask import make_response
    
    # Check permissions: Admin can view all, employee only their own
    if session.get('role') != 'admin' and session.get('user_id') != employee_id:
        flash('Access denied.', 'error')
        return redirect(url_for('employee_dashboard'))
    
    month = request.args.get('month')  # Format: YYYY-MM
    
    if not month:
        flash('Please select a month.', 'error')
        return redirect(url_for('view_employee_reports', employee_id=employee_id))
    
    conn = get_db_connection()
    
    # Get employee details
    employee = execute_query(conn, 'SELECT * FROM users WHERE id = ?', (employee_id,)).fetchone()
    if not employee:
        flash('Employee not found.', 'error')
        conn.close()
        return redirect(url_for('admin_dashboard'))
    
    # Get submissions for the specified month
    query = '''
        SELECT s.date as "Date", 
               s.submission_number as "Submission #",
               s.client_name as "Client",
               s.work_type as "Work Type",
               s.quantity as "Qty",
               s.work_text as "Description",
               s.created_at as "Submitted At"
        FROM submissions s 
        WHERE s.user_id = ? AND strftime('%Y-%m', s.date) = ?
        ORDER BY s.date ASC, s.created_at ASC
    '''
    # Replace placeholders for pd.read_sql_query
    _, q = get_db_info()
    df_query = query.replace('?', q)
    if q == '%s':
        df_query = df_query.replace("strftime('%Y-%m', s.date)", "TO_CHAR(s.date, 'YYYY-MM')")
        
    df = pd.read_sql_query(df_query, conn, params=(employee_id, month))
    conn.close()
    
    if df.empty:
        flash('No data found for the selected month.', 'info')
        return redirect(url_for('view_employee_reports', employee_id=employee_id))

    # Add Employee Name as second column
    df.insert(1, 'Employee Name', employee['name'])

    is_admin = session.get('role') == 'admin'
    
    if is_admin:
        # Calculate amount column
        df['Amount'] = df.apply(lambda row: calculate_submission_amount(row['Work Type'], row['Qty'], employee['employment_type']), axis=1)
        
        # Reorder columns to put Amount after Qty
        cols = list(df.columns)
        qty_idx = cols.index('Qty')
        cols.insert(qty_idx + 1, cols.pop(cols.index('Amount')))
        df = df[cols]

    # Create Excel file in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Monthly Report')
        
        # Access the worksheet to apply some basic formatting
        workbook = writer.book
        worksheet = writer.sheets['Monthly Report']
        
        # Adjust column widths
        for i, col in enumerate(df.columns):
            column_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
            worksheet.column_dimensions[chr(65 + i)].width = min(column_len, 50)
            
        # If freelancer and admin, add total row
        if is_admin and employee['employment_type'] == 'freelancer':
            total_amount = df['Amount'].sum()
            last_row = len(df) + 2
            qty_idx = list(df.columns).index('Qty')
            worksheet.cell(row=last_row, column=qty_idx + 1, value="TOTAL:")
            worksheet.cell(row=last_row, column=qty_idx + 2, value=total_amount)
            
            # Make total row bold
            from openpyxl.styles import Font
            worksheet.cell(row=last_row, column=qty_idx + 1).font = Font(bold=True)
            worksheet.cell(row=last_row, column=qty_idx + 2).font = Font(bold=True)

    output.seek(0)
    
    # Create response
    response = make_response(output.getvalue())
    filename = f"{employee['name'].replace(' ', '_')}_{month}_report.xlsx"
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    return response

@app.route('/admin/reports/download')
@admin_required
def download_filtered_reports():
    """Download filtered reports from admin dashboard as Excel"""
    from flask import make_response
    
    employee_filter = request.args.get('employee', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    employment_type_filter = request.args.get('employment_type', 'both')
    
    conn = get_db_connection()
    
    # Calculate amount
    filtered_submissions = []
    
    query = '''
        SELECT s.date as "Date", 
               u.name as "Employee Name",
               u.employment_type as "Type",
               s.client_category as "Category",
               s.client_name as "Client",
               s.work_type as "Work Type",
               s.quantity as "Qty",
               s.work_text as "Description",
               s.created_at as "Submitted At"
        FROM submissions s
        JOIN users u ON s.user_id = u.id
        WHERE 1=1
    '''
    
    params = []
    if employee_filter:
        query += ' AND (LOWER(u.name) LIKE LOWER(?) OR LOWER(s.employee_name) LIKE LOWER(?))'
        params.extend([f'%{employee_filter}%', f'%{employee_filter}%'])
    
    if start_date:
        query += ' AND s.date >= ?'
        params.append(start_date)
        
    if end_date:
        query += ' AND s.date <= ?'
        params.append(end_date)
        
    if employment_type_filter and employment_type_filter != 'both':
        query += ' AND u.employment_type = ?'
        params.append(employment_type_filter)
        
    query += ' ORDER BY s.date DESC, s.created_at DESC'
    
    df = pd.read_sql_query(query, conn, params=params)
    conn.close()
    
    if df.empty:
        return redirect(url_for('admin_dashboard'))

    # Add Amount column for freelancers
    # ... logic to calculate amount ...
    df['Amount'] = df.apply(lambda row: calculate_submission_amount(row['Work Type'], row['Qty'], row['Type']), axis=1)
    
    # Reorder columns: Date, Employee Name, Type, Category, Client, Work Type, Qty, Amount, Description, Submitted At
    desired_order = ['Date', 'Employee Name', 'Type', 'Category', 'Client', 'Work Type', 'Qty', 'Amount', 'Description', 'Submitted At']
    # Filter only columns that exist
    final_cols = [c for c in desired_order if c in df.columns]
    df = df[final_cols]

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Work Reports')
        
        # Auto-adjust columns
        worksheet = writer.sheets['Work Reports']
        for i, col in enumerate(df.columns):
            # Safe length calculation
            max_len = 0
            if not df[col].empty:
                max_len = df[col].astype(str).map(len).max()
            column_len = max(max_len, len(str(col))) + 2
            # Excel column letter (A, B, C...)
            col_letter = chr(65 + i)
            if i > 25: # For AA, AB etc if needed, though simple logic handles A-Z
                 col_letter = chr(65 + (i // 26) - 1) + chr(65 + (i % 26))
            
            try:
                worksheet.column_dimensions[col_letter].width = min(column_len, 50)
            except:
                pass # Fallback if column letter logic fails for many columns

    output.seek(0)
    
    response = make_response(output.getvalue())
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response.headers["Content-Disposition"] = f"attachment; filename=Work_Reports_Export_{timestamp}.xlsx"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    return response

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    """Handle user profile updates"""
    conn = get_db_connection()
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        password = request.form.get('password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        
        if not name:
            flash('Name is required.', 'error')
        elif password and password != confirm_password:
            flash('Passwords do not match.', 'error')
        else:
            try:
                if password:
                    hashed_password = generate_password_hash(password)
                    execute_query(conn, 'UPDATE users SET name = ?, password = ? WHERE id = ?',
                               (name, hashed_password, session['user_id']))
                else:
                    execute_query(conn, 'UPDATE users SET name = ? WHERE id = ?',
                               (name, session['user_id']))
                conn.commit()
                session['user_name'] = name
                flash('Profile updated successfully!', 'success')
            except Exception as e:
                flash(f'Error updating profile: {str(e)}', 'error')
                
    user = execute_query(conn, 'SELECT * FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    conn.close()
    
    return render_template('profile.html', user=user)

@app.route('/report/edit/<int:report_id>', methods=['GET', 'POST'])
@login_required
def edit_report(report_id):
    """Edit an existing work report"""
    conn = get_db_connection()
    report = execute_query(conn, 'SELECT * FROM submissions WHERE id = ?', (report_id,)).fetchone()
    
    if not report:
        conn.close()
        flash('Report not found.', 'error')
        return redirect(url_for('index'))
    
    report = dict(report)
        
    # Check permissions
    if session['role'] != 'admin' and report['user_id'] != session['user_id']:
        conn.close()
        flash('Access denied.', 'error')
        return redirect(url_for('index'))
        
    if request.method == 'POST':
        work_text = request.form.get('work_text', '').strip()
        client_category = request.form.get('client_category', '').strip()
        client_name = request.form.get('client_name', '').strip()
        other_client_name = request.form.get('other_client_name', '').strip()
        work_type = request.form.get('work_type', '').strip()
        quantity = request.form.get('quantity', '1')
        
        # Handle 'Other' client name
        if client_name in ['Others', 'Other'] and other_client_name:
            client_name = other_client_name
        
        if not work_text or not client_name or not work_type:
            flash('All fields marked with * are required.', 'error')
        else:
            file_path = report['file_path']
            if 'file' in request.files:
                file = request.files['file']
                if file and file.filename and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    filename = f"{session['user_id']}_{timestamp}_{filename}"
                    new_file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                    file.save(new_file_path)
                    
                    # Delete old file if exists
                    if file_path and os.path.exists(file_path):
                        try:
                            os.remove(file_path)
                        except:
                            pass
                    file_path = new_file_path
            
            try:
                execute_query(conn, 
                    '''UPDATE submissions 
                       SET work_text = ?, file_path = ?, client_category = ?, client_name = ?, work_type = ?, quantity = ? 
                       WHERE id = ?''',
                    (work_text, file_path, client_category, client_name, work_type, quantity, report_id)
                )
                conn.commit()
                flash('Report updated successfully!', 'success')
                return redirect(url_for('index'))
            except Exception as e:
                flash(f'Error updating report: {str(e)}', 'error')
    
    conn.close()
    return render_template('edit_report.html', report=report)

@app.route('/report/delete/<int:report_id>', methods=['POST'])
@login_required
def delete_report(report_id):
    """Delete a work report"""
    conn = get_db_connection()
    report = execute_query(conn, 'SELECT * FROM submissions WHERE id = ?', (report_id,)).fetchone()
    
    if not report:
        conn.close()
        flash('Report not found.', 'error')
        return redirect(url_for('index'))
        
    # Check permissions
    if session['role'] != 'admin' and report['user_id'] != session['user_id']:
        conn.close()
        flash('Access denied.', 'error')
        return redirect(url_for('index'))
        
    try:
        # Delete file if exists
        if report['file_path'] and os.path.exists(report['file_path']):
            try:
                os.remove(report['file_path'])
            except:
                pass
                
        execute_query(conn, 'DELETE FROM submissions WHERE id = ?', (report_id,))
        conn.commit()
        flash('Report deleted successfully!', 'success')
    except Exception as e:
        flash(f'Error deleting report: {str(e)}', 'error')
        
    conn.close()
    return redirect(url_for('index'))

@app.route('/admin/work-analysis')
@admin_required
def work_analysis():
    """Work Analysis - show how many of each work type each employee/freelancer did for a specific client"""
    conn = get_db_connection()

    # Filter parameters
    filter_client = request.args.get('client', '').strip()
    other_client = request.args.get('other_client', '').strip()
    filter_person = request.args.get('person', '').strip()
    filter_start = request.args.get('start_date', '').strip()
    filter_end = request.args.get('end_date', '').strip()

    if filter_client == 'Other' and other_client:
        filter_client = other_client

    # Define standard clients list based on corporate/political categories
    all_clients = ['IMC', 'Cornext', 'AIC', 'Yuvatha', 'Raksha', 'RMR', 'RSR', 'SG', 'JKR', 'Degala']

    # Get all employees and freelancers for the person dropdown
    all_persons_raw = execute_query(conn,
        "SELECT DISTINCT name, employment_type FROM users WHERE role = 'employee' ORDER BY name"
    ).fetchall()
    # Also get admin-submitted employee_name values (freelancer/other names added via admin form)
    extra_names_raw = execute_query(conn,
        "SELECT DISTINCT employee_name FROM submissions WHERE employee_name IS NOT NULL AND employee_name != '' ORDER BY employee_name"
    ).fetchall()
    extra_names = [r['employee_name'] for r in extra_names_raw]

    all_persons = [{'name': r['name'], 'employment_type': r['employment_type']} for r in all_persons_raw]
    # Add extra names that are not already in all_persons
    existing_names = {p['name'] for p in all_persons}
    for en in extra_names:
        if en not in existing_names:
            all_persons.append({'name': en, 'employment_type': 'freelancer'})

    # Build query for analysis
    query = '''
        SELECT 
            COALESCE(s.employee_name, u.name) as person_name,
            u.employment_type,
            s.client_name,
            s.work_type,
            SUM(CAST(s.quantity AS INTEGER)) as total_qty,
            COUNT(*) as entries
        FROM submissions s
        JOIN users u ON s.user_id = u.id
        WHERE 1=1
    '''
    params = []

    if filter_client:
        query += ' AND LOWER(s.client_name) = LOWER(?)'
        params.append(filter_client)

    if filter_person:
        query += ' AND (LOWER(u.name) = LOWER(?) OR LOWER(s.employee_name) = LOWER(?))'
        params.append(filter_person)
        params.append(filter_person)

    if filter_start:
        query += ' AND s.date >= ?'
        params.append(filter_start)

    if filter_end:
        query += ' AND s.date <= ?'
        params.append(filter_end)

    query += ' GROUP BY COALESCE(s.employee_name, u.name), u.employment_type, s.client_name, s.work_type ORDER BY person_name, s.client_name, s.work_type'

    raw_rows = [dict(r) for r in execute_query(conn, query, params).fetchall()]

    # Query for the Activity Streak chart (daily quantities per person)
    streak_query = '''
        SELECT 
            COALESCE(s.employee_name, u.name) as person_name,
            s.date as work_date,
            SUM(CAST(s.quantity AS INTEGER)) as daily_qty
        FROM submissions s
        JOIN users u ON s.user_id = u.id
        WHERE 1=1
    '''
    # We apply the same dynamic WHERE clauses and params as the first query.
    # We can reconstruct it simply:
    if filter_client:
        streak_query += ' AND LOWER(s.client_name) = LOWER(?)'
    if filter_person:
        streak_query += ' AND (LOWER(u.name) = LOWER(?) OR LOWER(s.employee_name) = LOWER(?))'
    if filter_start:
        streak_query += ' AND s.date >= ?'
    if filter_end:
        streak_query += ' AND s.date <= ?'

    streak_query += ' GROUP BY COALESCE(s.employee_name, u.name), s.date ORDER BY s.date ASC'
    raw_streak_results = execute_query(conn, streak_query, params).fetchall()
    raw_streak = format_streak_dates(raw_streak_results)

    conn.close()

    # Collect all unique work types from results
    all_work_types = sorted(set(r['work_type'] for r in raw_rows if r['work_type']))


    # Build pivot: keyed by (person_name, client_name) -> {work_type: total_qty, ...}
    from collections import defaultdict
    pivot = defaultdict(lambda: defaultdict(int))
    person_client_meta = {}

    for row in raw_rows:
        key = (row['person_name'], row['client_name'])
        pivot[key][row['work_type']] = row['total_qty']
        if key not in person_client_meta:
            person_client_meta[key] = {
                'person_name': row['person_name'],
                'employment_type': row['employment_type'],
                'client_name': row['client_name'],
            }

    # Sort keys: by person_name then client_name
    sorted_keys = sorted(pivot.keys(), key=lambda k: (k[0].lower(), k[1].lower() if k[1] else ''))

    # Build table rows
    table_rows = []
    for key in sorted_keys:
        meta = person_client_meta[key]
        work_counts = pivot[key]
        total = sum(work_counts.values())
        table_rows.append({
            'person_name': meta['person_name'],
            'employment_type': meta['employment_type'],
            'client_name': meta['client_name'],
            'work_counts': work_counts,
            'total': total,
        })

    return render_template(
        'work_analysis.html',
        all_clients=all_clients,
        all_persons=all_persons,
        filter_client=filter_client,
        filter_person=filter_person,
        filter_start=filter_start,
        filter_end=filter_end,
        all_work_types=all_work_types,
        table_rows=table_rows,
        raw_streak=raw_streak,
    )


if __name__ == '__main__':
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    
    # Database is initialized at the top level, no need to call it again here
    # unless we want to ensure it's fresh on every restart. 
    # But it's already called on line 166.
    
    print("\n" + "="*60)
    print("Work Report Management System")
    print("="*60)
    print("\nAdmin Login Details:")
    print("-" * 60)
    print("  Email: prashanth@iramediaconcepts.com")
    print("  Password: admin123")
    print("="*60 + "\n")
    
    # Run the Flask app
    app.run(debug=True, host='0.0.0.0', port=5001)
